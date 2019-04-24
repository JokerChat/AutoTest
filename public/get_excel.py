# -*- coding: utf-8 -*-
# @Author       :junjie    
# @Time         :2019/3/25 14:09
# @FileName     :get_excel.py
#IDE            :PyCharm
#封装Excel方法
from openpyxl import load_workbook
import os
from public.log import logger
mylog=logger('Excel模块').get_logger()
class getExcel(object):
    #初始化数据,传入文件路径和工作表
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name
    # 关闭Excel方法
    def close_excel(self):
        try:
            wb = load_workbook(self.file_path)
            wb.close()
            mylog.info("############关闭Excel成功############")
        except Exception as e:
            mylog.info("############关闭Excel出错:{}############".format(e))
    # 读取Excel方法
    def get_excel(self):

        #打开Excel文件
        wb=load_workbook(self.file_path)
        mylog.info("############打开Excel成功############")
        #打开对应的工作表
        sheet=wb[self.sheet_name]
        token=self.init_data()['token']
        mylog.info("############操作的工作表是{}############".format(self.sheet_name))
        #将测试数据存到一个列表里
        test_item=[]
        sql_item=[]
        try:
            if self.sheet_name=='test_data':
            #循环读取Excel表格,存到字典里,并对字典key进行赋值,列表嵌套字典,
                for i in range(2,sheet.max_row+1):
                    sub_data={}
                    sub_data['id']=int(sheet.cell(i, 1).value)
                    sub_data['name']=sheet.cell(i, 2).value
                    sub_data['url']=sheet.cell(i, 3).value
                    sub_data['method']=sheet.cell(i, 4).value
                    sub_data['code']=sheet.cell(i, 6).value
                    if sheet.cell(i, 5).value.find('$token')!=-1:
                        sub_data['data'] = sheet.cell(i, 5).value.replace('$token', token)
                    else:
                        sub_data['data'] = sheet.cell(i, 5).value#.replace('$token', token)
                    test_item.append(sub_data)
                mylog.info("############添加测试数据成功############")
                self.close_excel()
                #将测试数据返回
                return test_item
        except Exception as e:
            mylog.info("############关闭Excel出错:{}############".format(e))
        try:
            if self.sheet_name=='sql_data':
                for i in range(2,sheet.max_row+1):
                    sql_item.append({
                                    'id': int(sheet.cell(i, 1).value),
                                    'name':sheet.cell(i,2).value,
                                    'sql':sheet.cell(i, 3).value,
                                    'value':sheet.cell(i, 4).value,
                                        })
                mylog.info("############添加测试数据成功############")
                self.close_excel()
                return sql_item
        except Exception as e:
            mylog.info("############关闭Excel出错:{}############".format(e))

    #写入Excel方法
    def write_excel(self,row,column,value):
        '''
        :param row: 代表写入的行数
        :param column: 代表写入的列数
        :param value: 代表写入的值
        :return:
        '''
        try:
            wb=load_workbook(self.file_path)
            sheet=wb[self.sheet_name]
            mylog.info("############写入数据的工作表是{}############".format(self.sheet_name))
            #为Excel指定的行列写入指定的值
            mylog.info("############第{0}行第{1}列,写入值{2}############".format(row, column, value))
            sheet.cell(row,column).value=value
            mylog.info("############写入数据成功############")
            #写完结果后，保存Excel
            wb.save(self.file_path)
            # mylog.info("############第{0}行第{1}列,写入数据成功,写入数据:{2}############".format(row,column,value))
            self.close_excel()
        except Exception as e:
            mylog.info("############Excel写入出错:{}############".format(e))

    def init_data(self):
        wb=load_workbook(self.file_path)
        sheet=wb['update']
        token=sheet.cell(2,2).value
        return {'token':token}
        # test_data=self.get_excel()
        # for i in test_data:
        #     data = mysql.get_one(i['sql'])
        # for a in data.values():
        #     for row in range(len(test_data)):
        #         wb.write_excel(row + 2, 4, a)
    #更新Excel的数据
    def update_data(self):
        return self.get_excel()
if __name__=='__main__':
    file = os.path.abspath(os.path.join(os.path.dirname(__file__), "..")) + '\\test_file\\meiYe.xlsx'
    test_data=getExcel(file,'test_data').get_excel()
    for a in test_data:
        print(a)